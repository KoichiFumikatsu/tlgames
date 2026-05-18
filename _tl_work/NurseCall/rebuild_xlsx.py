"""Rebuild SystemText.dat replacing 'en' sheet content with ES translations.

Strategy:
- Load es_corpus.jsonl → dict id→target
- Load original .dat as openpyxl workbook (in-memory)
- Write ES text into 'en' sheet, preserving structure
- Change ID 23 display name from 'Fellow localize / English' to 'Español'
- Save as new file, then rename over original
"""
import io, sys, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from pathlib import Path
from openpyxl import load_workbook

DAT_ORIG = Path(r"c:/xampp/htdocs/tl/proyects Game TL/Unity/ナースコール警備員/NurseCall_Data/SystemText.dat")
CORPUS   = Path(r"c:/xampp/htdocs/tl/_tl_work/NurseCall/es_corpus.jsonl")
OUT_DAT  = Path(r"c:/xampp/htdocs/tl/_tl_work/NurseCall/SystemText_es.dat")

# Load translations
translations: dict[str, str] = {}
with open(CORPUS, encoding="utf-8") as f:
    for line in f:
        obj = json.loads(line)
        translations[str(obj["id"])] = obj["target"]

print(f"Loaded {len(translations)} translations")

# Load workbook (read-write mode — NOT read_only)
with open(DAT_ORIG, "rb") as fh:
    buf = io.BytesIO(fh.read())
wb = load_workbook(buf)

ws = wb["en"]
print(f"Sheet 'en': {ws.max_row} rows, {ws.max_column} cols")

patched = 0
skipped_comment = 0
skipped_empty = 0
skipped_no_tl = 0

for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
    if len(row) < 2:
        continue
    id_cell, text_cell = row[0], row[1]
    if id_cell.value is None:
        continue
    id_str = str(id_cell.value).strip()
    if not id_str:
        continue
    if id_str.startswith("//"):
        skipped_comment += 1
        continue

    if id_str in translations:
        tgt = translations[id_str]
        if tgt:
            text_cell.value = tgt
            patched += 1
        else:
            skipped_empty += 1
    else:
        skipped_no_tl += 1

print(f"Patched: {patched}  Skipped (comment): {skipped_comment}  "
      f"Skipped (empty tl): {skipped_empty}  Skipped (no tl): {skipped_no_tl}")

# Save to output path (use .xlsx extension for openpyxl, rename after)
tmp_xlsx = OUT_DAT.with_suffix(".xlsx")
wb.save(tmp_xlsx)
# Rename to .dat
if OUT_DAT.exists():
    OUT_DAT.unlink()
tmp_xlsx.rename(OUT_DAT)

print(f"Saved → {OUT_DAT}")
print(f"Size: {OUT_DAT.stat().st_size:,} bytes")
