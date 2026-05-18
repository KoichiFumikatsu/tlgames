"""Extract translatable strings from the 'en' sheet of SystemText.dat (XLSX).

Output: en_corpus.jsonl  — one JSON object per line: {"id": "...", "source": "..."}
Skips: empty strings, section comments (//), header rows (//ID)
"""
import io, sys, json, zipfile, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DAT = r"c:/xampp/htdocs/tl/proyects Game TL/Unity/ナースコール警備員/NurseCall_Data/SystemText.dat"
OUT = r"c:/xampp/htdocs/tl/_tl_work/NurseCall/en_corpus.jsonl"

from openpyxl import load_workbook

import io as _io
with open(DAT, "rb") as _fh:
    _buf = _io.BytesIO(_fh.read())
wb = load_workbook(_buf, read_only=True, data_only=True)
ws = wb["en"]

rows_written = 0
chars_total = 0

with open(OUT, "w", encoding="utf-8") as f:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        id_val, text_val = row[0], row[1]
        if id_val is None:
            continue
        id_str = str(id_val).strip()
        if not id_str:
            continue
        # Skip section comments and header rows
        if id_str.startswith("//"):
            continue
        text_str = str(text_val).strip() if text_val is not None else ""
        # Skip empty
        if not text_str:
            continue
        entry = {"id": id_str, "source": text_str}
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")
        rows_written += 1
        chars_total += len(text_str)

print(f"Extracted {rows_written} strings, {chars_total} chars → {OUT}")
wb.close()
