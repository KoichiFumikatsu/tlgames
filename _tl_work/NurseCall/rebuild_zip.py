"""Patch SystemText.dat at ZIP/XML level without openpyxl write.

Two operations:
1. EN sheet: replace all 681 translatable strings with ES translations
   (add new sharedString entries, update sheet2.xml references)
2. ID 23 in ALL sheets: replace with 'Español' directly in existing ss entries
   (so the language selector shows 'Español' regardless of active language)
"""
import io, sys, json, re, zipfile
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import xml.etree.ElementTree as ET

DAT_ORIG = r"c:/xampp/htdocs/tl/backups/Unity/NurseCall-SystemText-original-20260429/SystemText.dat"
CORPUS   = r"c:/xampp/htdocs/tl/_tl_work/NurseCall/es_corpus.jsonl"
OUT_DAT  = r"c:/xampp/htdocs/tl/_tl_work/NurseCall/SystemText_es3.dat"

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

# Load translations
translations: dict[str, str] = {}
with open(CORPUS, encoding="utf-8") as f:
    for line in f:
        obj = json.loads(line)
        translations[str(obj["id"])] = obj["target"]
print(f"Loaded {len(translations)} translations")

with zipfile.ZipFile(DAT_ORIG, "r") as zin:
    sheet_bytes = {
        "jp": zin.read("xl/worksheets/sheet1.xml"),
        "en": zin.read("xl/worksheets/sheet2.xml"),
        "sc": zin.read("xl/worksheets/sheet3.xml"),
        "tc": zin.read("xl/worksheets/sheet4.xml"),
        "ko": zin.read("xl/worksheets/sheet5.xml"),
    }
    ss_bytes = zin.read("xl/sharedStrings.xml")

def parse_id_to_ss(sheet_data: bytes) -> dict[str, str]:
    """Return {id_str: ss_index_str} for all numeric-ID rows with col-B shared string."""
    tree = ET.fromstring(sheet_data)
    sd = tree.find(f"{{{NS}}}sheetData")
    result = {}
    for row_el in sd.findall(f"{{{NS}}}row"):
        col_a_val = col_b_ss = None
        for c in row_el.findall(f"{{{NS}}}c"):
            ref = c.get("r", "")
            v_el = c.find(f"{{{NS}}}v")
            if v_el is None:
                continue
            if ref.startswith("A") and c.get("t") != "s":
                col_a_val = v_el.text
            elif ref.startswith("B") and c.get("t") == "s":
                col_b_ss = v_el.text
        if col_a_val and col_b_ss:
            result[col_a_val] = col_b_ss
    return result

# --- Map ID→ss for EN sheet ---
en_id_to_ss = parse_id_to_ss(sheet_bytes["en"])
print(f"EN sheet: {len(en_id_to_ss)} ID→ss pairs")

# --- Build ID-23 ss_index for all sheets ---
id23_ss: dict[str, int] = {}  # sheet_name -> ss_index for ID 23
for name, data in sheet_bytes.items():
    m = parse_id_to_ss(data)
    if "23" in m:
        id23_ss[name] = int(m["23"])
        print(f"  [{name}] ID 23 → ss_index {id23_ss[name]}")

# --- Parse sharedStrings.xml ---
ss_xml = ss_bytes.decode("utf-8")
sst_match = re.search(r'<sst[^>]+count="(\d+)"', ss_xml)
current_count = int(sst_match.group(1)) if sst_match else 3118
print(f"sharedStrings count: {current_count}")

# --- Operation 1: Directly replace ID-23 text in existing ss entries ---
# sharedStrings entries look like: <si><t>...</t></si> or <si><t xml:space="preserve">...</t></si>
# We identify by position (index). Build list of si positions.

# Find all <si>...</si> spans in order
si_pattern = re.compile(r"<si>.*?</si>", re.DOTALL)
si_spans = [(m.start(), m.end()) for m in si_pattern.finditer(ss_xml)]
print(f"Found {len(si_spans)} <si> entries in sharedStrings.xml")

# For each ID-23 ss_index across all sheets, replace the text
# We build a set of indices to replace
indices_to_replace = set(id23_ss.values())
print(f"ss indices to replace with 'Español': {sorted(indices_to_replace)}")

# Build patched ss_xml: replace text in those si entries
# We'll do this by reconstructing the string with replacements
ss_xml_patched = ss_xml
# Process in reverse order so offsets stay valid
for ss_idx in sorted(indices_to_replace, reverse=True):
    if ss_idx >= len(si_spans):
        print(f"  WARNING: ss_index {ss_idx} out of range!")
        continue
    start, end = si_spans[ss_idx]
    old_si = ss_xml_patched[start:end]
    # Replace the text content inside <t>...</t>
    new_si = re.sub(r"(<t[^>]*>)[^<]*(</t>)", r"\1Español\2", old_si)
    ss_xml_patched = ss_xml_patched[:start] + new_si + ss_xml_patched[end:]
    print(f"  Replaced ss[{ss_idx}]: {old_si!r} → {new_si!r}")

# --- Operation 2: Add new ss entries for EN translations ---
old_to_new: dict[str, str] = {}
new_si_entries: list[str] = []
new_idx = current_count

for id_str, es_text in translations.items():
    if id_str not in en_id_to_ss:
        continue
    old_ss = en_id_to_ss[id_str]
    es_escaped = (es_text
                  .replace("&", "&amp;")
                  .replace("<", "&lt;")
                  .replace(">", "&gt;")
                  .replace('"', "&quot;"))
    new_si_entries.append(f'<si><t xml:space="preserve">{es_escaped}</t></si>')
    old_to_new[old_ss] = str(new_idx)
    new_idx += 1

print(f"New EN ss entries: {len(new_si_entries)}")

# Update count in sharedStrings
total_count = new_idx
ss_xml_patched = re.sub(
    r'(<sst[^>]+count=")(\d+)"',
    lambda m: f'{m.group(1)}{total_count}"',
    ss_xml_patched
)
# Append new entries before </sst>
new_entries_block = "\r\n".join(new_si_entries)
ss_xml_patched = ss_xml_patched.rstrip()
if ss_xml_patched.endswith("</sst>"):
    ss_xml_patched = ss_xml_patched[:-6] + new_entries_block + "</sst>"

# --- Patch sheet2.xml (EN) references ---
sheet2_xml = sheet_bytes["en"].decode("utf-8")

def patch_sheet2(xml_str: str, old_to_new: dict[str, str]) -> str:
    def replace_b_cell(m: re.Match) -> str:
        full = m.group(0)
        v_val = m.group(1)
        if v_val in old_to_new:
            return full.replace(f"<v>{v_val}</v>", f"<v>{old_to_new[v_val]}</v>", 1)
        return full
    pattern = r'<c r="B\d+"[^>]*t="s"[^>]*><v>(\d+)</v></c>'
    return re.sub(pattern, replace_b_cell, xml_str)

new_sheet2_xml = patch_sheet2(sheet2_xml, old_to_new)

replaced_count = sum(1 for new in old_to_new.values() if f"<v>{new}</v>" in new_sheet2_xml)
print(f"Verified EN sheet replacements: {replaced_count}/{len(old_to_new)}")

# --- Repack ZIP ---
new_ss_bytes     = ss_xml_patched.encode("utf-8")
new_sheet2_bytes = new_sheet2_xml.encode("utf-8")

with zipfile.ZipFile(DAT_ORIG, "r") as zin:
    with zipfile.ZipFile(OUT_DAT, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == "xl/sharedStrings.xml":
                zout.writestr(item, new_ss_bytes)
            elif item.filename == "xl/worksheets/sheet2.xml":
                zout.writestr(item, new_sheet2_bytes)
            else:
                zout.writestr(item, zin.read(item.filename))

import os
print(f"\nSaved → {OUT_DAT}  ({os.path.getsize(OUT_DAT):,} bytes)")

# Quick validation
from openpyxl import load_workbook
with open(OUT_DAT, "rb") as fh:
    buf = io.BytesIO(fh.read())
wb = load_workbook(buf, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
for sheet_name in ["jp", "en", "sc", "tc", "ko"]:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=4, min_col=1, max_col=2, values_only=True):
        if row[0] is not None and str(row[0]).strip() == "23":
            print(f"  [{sheet_name}] ID 23: {row[1]!r}")
            break
wb.close()
